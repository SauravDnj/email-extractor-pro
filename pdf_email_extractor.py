import re
import random
import json
from datetime import datetime
import PyPDF2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

class PDFEmailExtractor:
    """
    PDF Email Extractor System
    - Extracts emails from PDF files
    - Randomly selects 200 unique emails
    - No duplicate selection across multiple runs
    - Exports to Excel with beautiful formatting
    """
    
    def __init__(self, history_file='pdf_email_history.json'):
        self.history_file = history_file
        self.all_emails = set()
        self.selected_emails = []
        self.previously_selected = self.load_history()
        self.source_pdf = None
        
    def load_history(self):
        """Load previously selected emails from history file"""
        try:
            with open(self.history_file, 'r') as f:
                return set(json.load(f))
        except FileNotFoundError:
            return set()
    
    def save_history(self):
        """Save selected emails to history file"""
        all_history = self.previously_selected.union(set(self.selected_emails))
        with open(self.history_file, 'w') as f:
            json.dump(list(all_history), f, indent=2)
        print(f"✅ History saved: {len(all_history)} total emails tracked")
    
    def extract_text_from_pdf(self, pdf_path):
        """Extract all text from PDF file"""
        try:
            print(f"\n📄 Reading PDF: {pdf_path}")
            self.source_pdf = pdf_path
            
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                total_pages = len(pdf_reader.pages)
                print(f"   Total pages: {total_pages}")
                
                text = ""
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    print(f"   Processing page {page_num}/{total_pages}...", end='\r')
                    text += page.extract_text() + "\n"
                
                print(f"\n   ✓ Successfully extracted text from {total_pages} pages")
                return text
                
        except FileNotFoundError:
            print(f"❌ Error: PDF file not found: {pdf_path}")
            return ""
        except Exception as e:
            print(f"❌ Error reading PDF: {e}")
            return ""
    
    def extract_emails_from_text(self, text):
        """Extract all email addresses from text"""
        # Comprehensive email regex pattern
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        emails = re.findall(email_pattern, text)
        
        # Convert to lowercase and add to set (removes duplicates)
        unique_emails = set(email.lower() for email in emails)
        
        return unique_emails
    
    def extract_emails_from_pdf(self, pdf_path):
        """Extract all emails from PDF file"""
        text = self.extract_text_from_pdf(pdf_path)
        
        if not text:
            return set()
        
        print(f"\n🔍 Searching for email addresses...")
        emails = self.extract_emails_from_text(text)
        
        # Add to all_emails set
        self.all_emails.update(emails)
        
        print(f"   ✓ Found {len(emails)} unique emails in this PDF")
        print(f"   ✓ Total unique emails: {len(self.all_emails)}")
        
        return emails
    
    def extract_from_multiple_pdfs(self, pdf_files):
        """Extract emails from multiple PDF files"""
        print(f"\n📚 Processing {len(pdf_files)} PDF files...")
        
        for pdf_file in pdf_files:
            self.extract_emails_from_pdf(pdf_file)
        
        return self.all_emails
    
    def get_available_emails(self):
        """Get emails that haven't been selected before"""
        available = list(self.all_emails - self.previously_selected)
        return available
    
    def select_random_200(self):
        """Randomly select 200 emails (or less if not available)"""
        available = self.get_available_emails()
        
        print(f"\n🎲 Selecting random emails...")
        print(f"   Available emails: {len(available)}")
        print(f"   Previously selected: {len(self.previously_selected)}")
        
        if len(available) == 0:
            print("\n⚠️ WARNING: No new emails available!")
            print("   All emails have been previously selected.")
            print("\n   Options:")
            print("   1. Add more PDF files with new emails")
            print("   2. Reset history using reset_history()")
            return []
        
        count = min(200, len(available))
        
        if len(available) < 200:
            print(f"   ⚠️ Only {len(available)} new emails available")
            print(f"   Selecting all {count} available emails")
        else:
            print(f"   ✓ Selecting 200 random emails")
        
        # Randomly select emails
        self.selected_emails = random.sample(available, count)
        
        print(f"   ✓ Selected {len(self.selected_emails)} emails")
        
        return self.selected_emails
    
    def export_to_excel(self, filename='pdf_emails_200.xlsx'):
        """Export selected emails to Excel with beautiful formatting"""
        
        if not self.selected_emails:
            print("❌ No emails to export! Please select emails first.")
            return None
        
        print(f"\n💾 Creating Excel file: {filename}")
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Random 200 Emails"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        border_style = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Add title row
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"📧 Random 200 Emails Extracted from PDF"
        title_cell.font = Font(bold=True, size=14, color="4472C4")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Add headers
        headers = ['#', 'Email Address', 'Domain', 'Provider Type', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        
        # Add data
        for idx, email in enumerate(self.selected_emails, 1):
            # Extract domain
            domain = email.split('@')[1] if '@' in email else 'N/A'
            
            # Determine provider type
            if 'gmail' in domain:
                provider = 'Gmail'
            elif 'yahoo' in domain:
                provider = 'Yahoo'
            elif 'outlook' in domain or 'hotmail' in domain:
                provider = 'Outlook'
            elif any(x in domain for x in ['.edu', '.ac.', 'university']):
                provider = 'Educational'
            elif any(x in domain for x in ['.gov', 'government']):
                provider = 'Government'
            else:
                provider = 'Corporate/Other'
            
            row = idx + 2
            
            # Row number
            cell = ws.cell(row=row, column=1, value=idx)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border_style
            
            # Email
            cell = ws.cell(row=row, column=2, value=email)
            cell.border = border_style
            
            # Domain
            cell = ws.cell(row=row, column=3, value=domain)
            cell.border = border_style
            
            # Provider Type
            cell = ws.cell(row=row, column=4, value=provider)
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center')
            
            # Status
            cell = ws.cell(row=row, column=5, value="New")
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.font = Font(color="006100", bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border_style
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        
        # Add Summary Sheet
        summary_ws = wb.create_sheet("Summary & Statistics")
        
        # Summary title
        summary_ws.merge_cells('A1:B1')
        summary_ws['A1'] = "📊 Email Extraction Summary"
        summary_ws['A1'].font = Font(bold=True, size=14, color="4472C4")
        summary_ws['A1'].alignment = Alignment(horizontal='center')
        summary_ws.row_dimensions[1].height = 30
        
        # Summary data
        summary_data = [
            ['Source PDF:', Path(self.source_pdf).name if self.source_pdf else 'Multiple PDFs'],
            ['Total Unique Emails Found:', len(self.all_emails)],
            ['Previously Selected (History):', len(self.previously_selected)],
            ['Available for Selection:', len(self.get_available_emails()) + len(self.selected_emails)],
            ['Emails in This Export:', len(self.selected_emails)],
            ['Extraction Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['Email Provider Breakdown:', '']
        ]
        
        # Count providers
        provider_counts = {}
        for email in self.selected_emails:
            domain = email.split('@')[1] if '@' in email else 'Unknown'
            if 'gmail' in domain:
                provider = 'Gmail'
            elif 'yahoo' in domain:
                provider = 'Yahoo'
            elif 'outlook' in domain or 'hotmail' in domain:
                provider = 'Outlook'
            elif any(x in domain for x in ['.edu', '.ac.', 'university']):
                provider = 'Educational'
            elif any(x in domain for x in ['.gov', 'government']):
                provider = 'Government'
            else:
                provider = 'Corporate/Other'
            
            provider_counts[provider] = provider_counts.get(provider, 0) + 1
        
        for provider, count in sorted(provider_counts.items(), key=lambda x: x[1], reverse=True):
            summary_data.append([f'  • {provider}:', count])
        
        # Write summary data
        for idx, (label, value) in enumerate(summary_data, 3):
            summary_ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=idx, column=2, value=value)
        
        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 30
        
        # Save workbook
        wb.save(filename)
        print(f"   ✓ Excel file created successfully!")
        print(f"   ✓ Location: {filename}")
        print(f"   ✓ Total emails exported: {len(self.selected_emails)}")
        
        return filename
    
    def get_statistics(self):
        """Get detailed statistics"""
        available = self.get_available_emails()
        
        stats = {
            'total_emails_found': len(self.all_emails),
            'previously_selected': len(self.previously_selected),
            'available_for_selection': len(available),
            'selected_this_time': len(self.selected_emails),
            'selection_percentage': f"{(len(self.selected_emails) / len(self.all_emails) * 100):.1f}%" if self.all_emails else "0%"
        }
        
        return stats
    
    def reset_history(self):
        """Reset the selection history"""
        self.previously_selected = set()
        try:
            import os
            if os.path.exists(self.history_file):
                os.remove(self.history_file)
            print("✅ History reset successfully!")
            print("   All emails are now available for selection again.")
        except Exception as e:
            print(f"❌ Error resetting history: {e}")
    
    def display_preview(self, count=10):
        """Display preview of selected emails"""
        if not self.selected_emails:
            print("No emails selected yet!")
            return
        
        print(f"\n📋 Preview of selected emails (first {count}):")
        print("=" * 60)
        
        for i, email in enumerate(self.selected_emails[:count], 1):
            domain = email.split('@')[1] if '@' in email else 'N/A'
            print(f"   {i:3d}. {email:40s} (@{domain})")
        
        if len(self.selected_emails) > count:
            print(f"   ... and {len(self.selected_emails) - count} more emails")
        
        print("=" * 60)


# Main execution function
def main():
    """Main function to demonstrate PDF email extraction"""
    
    print("=" * 70)
    print(" 📧 PDF EMAIL EXTRACTOR - Random 200 Unique Emails")
    print("=" * 70)
    print()
    
    # Initialize extractor
    extractor = PDFEmailExtractor()
    
    # Example: Extract from PDF file
    # Replace 'your_document.pdf' with your actual PDF file path
    pdf_file = input("Enter PDF file path (or press Enter for demo): ").strip()
    
    if not pdf_file:
        print("\n⚠️ No PDF file provided. Using demo mode.")
        print("\nTo use this system:")
        print("1. Place your PDF file in the same directory")
        print("2. Run the script again")
        print("3. Enter the PDF filename when prompted")
        return
    
    # Extract emails from PDF
    extractor.extract_emails_from_pdf(pdf_file)
    
    # Show statistics
    print("\n" + "=" * 70)
    print(" 📊 STATISTICS")
    print("=" * 70)
    stats = extractor.get_statistics()
    for key, value in stats.items():
        print(f"   {key.replace('_', ' ').title():30s}: {value}")
    
    # Select random 200 emails
    print("\n" + "=" * 70)
    selected = extractor.select_random_200()
    
    if selected:
        # Display preview
        extractor.display_preview(10)
        
        # Export to Excel
        print("\n" + "=" * 70)
        excel_file = extractor.export_to_excel()
        
        # Save history
        extractor.save_history()
        
        print("\n" + "=" * 70)
        print(" ✅ PROCESS COMPLETED SUCCESSFULLY!")
        print("=" * 70)
        print(f"\n📁 Excel File: {excel_file}")
        print(f"📊 Emails Exported: {len(selected)}")
        print(f"\n💡 TIP: Run again to get another 200 different emails!")
        print("   (Previously selected emails will be excluded)")
        print("=" * 70)
    else:
        print("\n" + "=" * 70)
        print(" ⚠️ NO EMAILS AVAILABLE")
        print("=" * 70)


if __name__ == "__main__":
    main()
